"""Read/write the ENACOM_DDJJ_Tracker_OsvaldoDemo.xlsx tracker.

Single source of truth for which DDJJ are pending, in progress, or finalized.
"""
from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

SHEET = "Tracker DDJJ"

# Column indices in the Tracker DDJJ sheet (1-based)
COL_NUM = 1
COL_SERVICIO = 2
COL_ANIO = 3
COL_PERIODO_NUM = 4
COL_PERIODO_NOMBRE = 5
COL_ESTADO = 6
COL_CARPETA = 7
COL_FECHA = 8
COL_FUNDAMENTO = 9
COL_NOTAS = 10


@dataclass
class DdjjRow:
    row: int
    num: int
    servicio: str            # "TCFV" | "SU-M" | "SU-T"
    anio: int
    periodo_num: int         # 1..12
    periodo_nombre: str      # "Enero", ...
    estado: str              # "Pendiente" | "En curso" | "Enviada" | "Validada" | "Finalizada" | "Rechazada"
    carpeta_tecnica: str | None
    fecha_presentacion: str | None
    fundamento: str | None
    notas: str | None

    def as_dict(self) -> dict:
        d = asdict(self)
        return d


class Tracker:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Tracker not found: {self.path}")

    def _load(self):
        return load_workbook(self.path)

    def list_all(self) -> list[DdjjRow]:
        wb = self._load()
        ws = wb[SHEET]
        rows: list[DdjjRow] = []
        for row in range(2, ws.max_row + 1):
            num = ws.cell(row=row, column=COL_NUM).value
            if num is None:
                break
            rows.append(DdjjRow(
                row=row,
                num=int(num),
                servicio=str(ws.cell(row=row, column=COL_SERVICIO).value),
                anio=int(ws.cell(row=row, column=COL_ANIO).value),
                periodo_num=int(ws.cell(row=row, column=COL_PERIODO_NUM).value),
                periodo_nombre=str(ws.cell(row=row, column=COL_PERIODO_NOMBRE).value),
                estado=str(ws.cell(row=row, column=COL_ESTADO).value or "Pendiente"),
                carpeta_tecnica=ws.cell(row=row, column=COL_CARPETA).value,
                fecha_presentacion=_fmt_date(ws.cell(row=row, column=COL_FECHA).value),
                fundamento=ws.cell(row=row, column=COL_FUNDAMENTO).value,
                notas=ws.cell(row=row, column=COL_NOTAS).value,
            ))
        return rows

    def list_pending(self) -> list[DdjjRow]:
        return [r for r in self.list_all() if r.estado.lower() == "pendiente"]

    def find(self, servicio: str, anio: int, periodo_num: int) -> DdjjRow | None:
        for r in self.list_all():
            if r.servicio == servicio and r.anio == anio and r.periodo_num == periodo_num:
                return r
        return None

    def update(
        self,
        servicio: str,
        anio: int,
        periodo_num: int,
        *,
        estado: str | None = None,
        carpeta_tecnica: str | None = None,
        fecha_presentacion: str | None = None,
        fundamento: str | None = None,
        notas: str | None = None,
    ) -> DdjjRow:
        wb = self._load()
        ws = wb[SHEET]
        row = self._find_row(ws, servicio, anio, periodo_num)
        if row is None:
            raise ValueError(f"Row not found for {servicio} {anio}-{periodo_num:02d}")
        if estado is not None:
            ws.cell(row=row, column=COL_ESTADO, value=estado)
        if carpeta_tecnica is not None:
            ws.cell(row=row, column=COL_CARPETA, value=carpeta_tecnica)
        if fecha_presentacion is not None:
            ws.cell(row=row, column=COL_FECHA, value=fecha_presentacion)
        if fundamento is not None:
            ws.cell(row=row, column=COL_FUNDAMENTO, value=fundamento)
        if notas is not None:
            ws.cell(row=row, column=COL_NOTAS, value=notas)
        wb.save(self.path)
        return self.find(servicio, anio, periodo_num)  # type: ignore[return-value]

    def summary(self) -> dict:
        rows = self.list_all()
        total = len(rows)
        by_status: dict[str, int] = {}
        for r in rows:
            by_status[r.estado] = by_status.get(r.estado, 0) + 1
        return {"total": total, "by_status": by_status}

    @staticmethod
    def _find_row(ws, servicio: str, anio: int, periodo_num: int) -> int | None:
        for row in range(2, ws.max_row + 1):
            if (
                ws.cell(row=row, column=COL_SERVICIO).value == servicio
                and ws.cell(row=row, column=COL_ANIO).value == anio
                and ws.cell(row=row, column=COL_PERIODO_NUM).value == periodo_num
            ):
                return row
        return None


def _fmt_date(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    return str(v)
