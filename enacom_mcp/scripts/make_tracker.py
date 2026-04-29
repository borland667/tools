"""Generate an Excel tracker template for ENACOM DDJJ.

Usage:
    python -m enacom_mcp.scripts.make_tracker \
        --cuit 20-XXXXXXXXX-X \
        --razon-social "NOMBRE APELLIDO" \
        --tcfv 2023-01:2024-12,2025-06:2025-12 \
        --su-m 2023-01:2024-12,2025-07:2025-12 \
        --output ./tracker.xlsx

Period spec: comma-separated ranges, each YYYY-MM:YYYY-MM (inclusive on both ends).

The tracker has three sheets:
  • "Tracker DDJJ"     — one row per (servicio, año, mes), with status dropdown
  • "Resumen"          — auto-counts of pending / sent / finalized
  • "DJ FINALIZADAS"   — pre-formatted to copy-paste once filing is done
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]
SERVICIOS = ("TCFV", "SU-M", "SU-T")


@dataclass(frozen=True)
class Period:
    servicio: str
    anio: int
    mes: int

    @property
    def mes_nombre(self) -> str:
        return MESES[self.mes - 1]


def parse_range_spec(spec: str, servicio: str) -> list[Period]:
    """Parse "2023-01:2024-12,2025-06:2025-12" into a list of Period."""
    out: list[Period] = []
    for chunk in spec.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        m = re.fullmatch(r"(\d{4})-(\d{1,2}):(\d{4})-(\d{1,2})", chunk)
        if not m:
            raise ValueError(f"Invalid range '{chunk}'. Expected YYYY-MM:YYYY-MM")
        y1, m1, y2, m2 = map(int, m.groups())
        if (y1, m1) > (y2, m2):
            raise ValueError(f"Inverted range '{chunk}'")
        y, mo = y1, m1
        while (y, mo) <= (y2, m2):
            if not 1 <= mo <= 12:
                raise ValueError(f"Bad month in '{chunk}'")
            out.append(Period(servicio, y, mo))
            mo += 1
            if mo == 13:
                mo, y = 1, y + 1
    return out


def build_tracker(cuit: str, razon_social: str, periods: list[Period], output: Path) -> int:
    wb = Workbook()

    # ---------- Sheet 1: Tracker ----------
    ws = wb.active
    ws.title = "Tracker DDJJ"

    headers = ["#", "Servicio", "Año", "Período", "Mes/Trim", "Estado",
               "Nº Carpeta Técnica", "Fecha presentación", "Fundamento usado", "Notas"]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E78")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    # Sort by year, service, month for predictable order.
    periods_sorted = sorted(periods, key=lambda p: (p.anio, p.servicio, p.mes))

    base_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for idx, p in enumerate(periods_sorted, start=1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=p.servicio)
        ws.cell(row=row, column=3, value=p.anio).number_format = "0"
        ws.cell(row=row, column=4, value=p.mes)
        ws.cell(row=row, column=5, value=p.mes_nombre)
        ws.cell(row=row, column=6, value="Pendiente")
        for col in (7, 8, 9, 10):
            ws.cell(row=row, column=col, value="")
        for col in range(1, 11):
            c = ws.cell(row=row, column=col)
            c.font = base_font
            c.border = border
            c.alignment = center if col in (1, 3, 4, 6) else left

    n = len(periods_sorted)
    last_row = n + 1
    widths = {1: 5, 2: 9, 3: 7, 4: 9, 5: 13, 6: 13, 7: 24, 8: 14, 9: 18, 10: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{last_row}"

    green = PatternFill("solid", start_color="C6EFCE")
    yellow = PatternFill("solid", start_color="FFEB9C")
    red = PatternFill("solid", start_color="FFC7CE")
    rng = f"F2:F{last_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Finalizada"'], fill=green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"En curso"'], fill=yellow))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pendiente"'], fill=red))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Rechazada"'], fill=red))

    dv_estado = DataValidation(
        type="list",
        formula1='"Pendiente,En curso,Enviada,Validada,Finalizada,Rechazada"',
        allow_blank=True,
    )
    dv_estado.add(rng)
    ws.add_data_validation(dv_estado)

    dv_serv = DataValidation(type="list", formula1=f'"{",".join(SERVICIOS)}"', allow_blank=False)
    dv_serv.add(f"B2:B{last_row}")
    ws.add_data_validation(dv_serv)

    # ---------- Sheet 2: Resumen ----------
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = f"Resumen DDJJ ENACOM — CUIT {cuit}"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    ws2.merge_cells("A1:C1")
    ws2["A3"] = "Total a presentar:"
    ws2["B3"] = f"=COUNTA('Tracker DDJJ'!A2:A{last_row})"
    for i, label in enumerate(
        ["Pendientes", "En curso", "Enviadas", "Validadas", "Finalizadas", "Rechazadas"], start=4
    ):
        ws2[f"A{i}"] = f"{label}:"
        ws2[f"B{i}"] = f'=COUNTIF(\'Tracker DDJJ\'!F2:F{last_row},"{label[:-1] if label.endswith("s") else label}")'
    ws2["A11"] = "% completado:"
    ws2["B11"] = f"=B8/B3"
    ws2["B11"].number_format = "0.0%"

    ws2["A13"] = "Por servicio"
    ws2["A13"].font = Font(name="Arial", bold=True, size=12)
    row = 14
    for serv in sorted({p.servicio for p in periods_sorted}):
        ws2[f"A{row}"] = f"{serv} total:"
        ws2[f"B{row}"] = f'=COUNTIF(\'Tracker DDJJ\'!B2:B{last_row},"{serv}")'
        row += 1
        ws2[f"A{row}"] = f"{serv} finalizadas:"
        ws2[f"B{row}"] = (
            f'=COUNTIFS(\'Tracker DDJJ\'!B2:B{last_row},"{serv}",'
            f'\'Tracker DDJJ\'!F2:F{last_row},"Finalizada")'
        )
        row += 1

    for r in range(3, row):
        for col in ("A", "B"):
            ws2[f"{col}{r}"].font = Font(name="Arial", size=11)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12

    # ---------- Sheet 3: DJ FINALIZADAS (export-ready) ----------
    ws3 = wb.create_sheet("DJ FINALIZADAS")
    ws3["A1"] = "Hoja exportable — pegar al mail final a ENACOM una vez presentadas todas las DDJJ"
    ws3["A1"].font = Font(name="Arial", bold=True, italic=True, size=10, color="C00000")
    ws3.merge_cells("A1:G1")
    headers3 = ["CUIT", "Razón Social", "Servicio", "Año", "Período", "Nº Carpeta Técnica", "Fecha presentación"]
    for col, h in enumerate(headers3, start=1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    for idx in range(n):
        src_row = idx + 2
        dst_row = idx + 4
        ws3.cell(row=dst_row, column=1, value=cuit)
        ws3.cell(row=dst_row, column=2, value=razon_social)
        ws3.cell(row=dst_row, column=3, value=f"='Tracker DDJJ'!B{src_row}")
        ws3.cell(row=dst_row, column=4, value=f"='Tracker DDJJ'!C{src_row}")
        ws3.cell(row=dst_row, column=5, value=f"='Tracker DDJJ'!E{src_row}")
        ws3.cell(row=dst_row, column=6, value=f"='Tracker DDJJ'!G{src_row}")
        ws3.cell(row=dst_row, column=7, value=f"='Tracker DDJJ'!H{src_row}")
        for col in range(1, 8):
            c = ws3.cell(row=dst_row, column=col)
            c.font = base_font
            c.border = border
            c.alignment = center if col in (1, 3, 4, 5) else left
        ws3.cell(row=dst_row, column=4).number_format = "0"
        ws3.cell(row=dst_row, column=7).number_format = "dd/mm/yyyy"

    widths3 = {1: 18, 2: 24, 3: 9, 4: 7, 5: 13, 6: 24, 7: 18}
    for col, w in widths3.items():
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = "A4"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return n


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        prog="make_tracker",
        description="Generate an Excel tracker for ENACOM DDJJ filings.",
    )
    p.add_argument("--cuit", required=True, help="CUIT (e.g. 20-XXXXXXXXX-X)")
    p.add_argument("--razon-social", required=True, help="Razón social / nombre del titular")
    p.add_argument("--tcfv", default="", help="Period ranges for TCFV: 'YYYY-MM:YYYY-MM,...'")
    p.add_argument("--su-m", default="", help="Period ranges for SU-M (mensual): 'YYYY-MM:YYYY-MM,...'")
    p.add_argument("--su-t", default="", help="Period ranges for SU-T (trimestral): 'YYYY-MM:YYYY-MM,...'")
    p.add_argument("--output", "-o", required=True, type=Path, help="Output .xlsx path")
    args = p.parse_args(argv)

    periods: list[Period] = []
    if args.tcfv:
        periods += parse_range_spec(args.tcfv, "TCFV")
    if args.su_m:
        periods += parse_range_spec(args.su_m, "SU-M")
    if args.su_t:
        periods += parse_range_spec(args.su_t, "SU-T")

    if not periods:
        print("error: at least one of --tcfv / --su-m / --su-t is required", file=sys.stderr)
        return 2

    n = build_tracker(args.cuit, args.razon_social, periods, args.output)
    print(f"Wrote {n} periods to {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
